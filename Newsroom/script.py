#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import time
from collections import Counter
from pathlib import Path
from urllib.parse import urlparse

import openpyxl
import requests


YEAR_RE = re.compile(r"(?<!\d)(20(?:1[0-9]|2[0-9]|30))(?!\d)")
YEAR_RANGE_RE = re.compile(
    r"(?<!\d)(20(?:1[0-9]|2[0-9]|30))\s*[–-]\s*(20(?:1[0-9]|2[0-9]|30))(?!\d)"
)
NUMERIC_DATE_RE = re.compile(
    r"(?<!\d)(\d{1,2})/(\d{1,2})/(20(?:1[0-9]|2[0-9]|30))(?!\d)"
)
TEXT_DATE_DAY_FIRST_RE = re.compile(
    r"(?<!\d)(\d{1,2})\s+"
    r"(january|february|march|april|may|june|july|august|september|october|november|december)"
    r"\s+(20(?:1[0-9]|2[0-9]|30))(?!\d)",
    re.IGNORECASE,
)
TEXT_DATE_MONTH_FIRST_RE = re.compile(
    r"(?<!\d)"
    r"(january|february|march|april|may|june|july|august|september|october|november|december)"
    r"\s+(\d{1,2}),?\s+(20(?:1[0-9]|2[0-9]|30))(?!\d)",
    re.IGNORECASE,
)


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    return "".join(ch.lower() for ch in str(value) if ch.isalnum())


def extract_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def infer_publication_timeframe(url: str, title: str) -> str | None:
    text = f"{url} {title}".strip()
    if not text:
        return None

    # Keep explicit ranges such as 2023-2024 when present.
    for match in YEAR_RANGE_RE.finditer(text):
        start, end = match.groups()
        if int(end) >= int(start):
            return f"{start}-{end}"

    years = [int(y) for y in YEAR_RE.findall(text)]
    if not years:
        return None
    return str(max(years))


def infer_topic(existing: str, section: str, page_type: str) -> str | None:
    if existing:
        return existing
    if section:
        return section
    if page_type:
        return page_type
    return None


def build_rjina_url(page_url: str) -> str:
    cleaned = page_url.split("://", 1)[-1]
    return f"https://r.jina.ai/http://{cleaned}"


def extract_explicit_date_years(markdown_text: str) -> list[int]:
    text = markdown_text.split("Markdown Content:", 1)[-1][:15000]
    years: list[int] = []

    for day_or_month, month_or_day, year in NUMERIC_DATE_RE.findall(text):
        first = int(day_or_month)
        second = int(month_or_day)
        year_value = int(year)
        if 1 <= first <= 31 and 1 <= second <= 31:
            years.append(year_value)

    for _, _, year in TEXT_DATE_DAY_FIRST_RE.findall(text):
        years.append(int(year))

    for _, _, year in TEXT_DATE_MONTH_FIRST_RE.findall(text):
        years.append(int(year))

    return years


def fetch_rjina_year(
    page_url: str, timeout_seconds: int, max_retries: int
) -> tuple[str | None, int]:
    parsed = urlparse(page_url)
    if parsed.path in {"", "/"}:
        return None, 0

    url = build_rjina_url(page_url)
    for _ in range(max_retries + 1):
        try:
            response = requests.get(url, timeout=timeout_seconds)
        except requests.RequestException:
            return None, 0

        if response.status_code == 429:
            retry_after = 4
            try:
                body = response.json()
                retry_after = int(body.get("retryAfter", retry_after))
            except ValueError:
                pass
            time.sleep(max(1, retry_after))
            continue

        if response.status_code != 200:
            return None, 0

        years = extract_explicit_date_years(response.text)
        if not years:
            return None, 0

        top_year, top_count = Counter(years).most_common(1)[0]
        return str(top_year), top_count

    return None, 0


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Backfill publication timeframe/topic values in a newsroom workbook."
    )
    parser.add_argument("input_file", type=Path)
    parser.add_argument(
        "--output-file",
        type=Path,
        default=Path("RLX_newsroom_pages_nov2025_filled.xlsx"),
    )
    parser.add_argument("--sheet-name", default=None)
    parser.add_argument(
        "--use-rjina",
        action="store_true",
        help="Try extracting explicit dates from page content through r.jina.ai for unresolved rows.",
    )
    parser.add_argument(
        "--rjina-timeout-seconds",
        type=int,
        default=35,
        help="Per-request timeout used with --use-rjina.",
    )
    parser.add_argument(
        "--rjina-min-interval-seconds",
        type=float,
        default=3.2,
        help="Minimum wait between r.jina.ai requests to reduce rate-limit hits.",
    )
    parser.add_argument(
        "--rjina-max-retries",
        type=int,
        default=4,
        help="Retries for r.jina.ai requests when rate-limited.",
    )
    args = parser.parse_args()

    wb = openpyxl.load_workbook(args.input_file)
    ws = wb[args.sheet_name] if args.sheet_name else wb[wb.sheetnames[0]]

    header_cells = list(next(ws.iter_rows(min_row=1, max_row=1)))
    headers = [extract_text(cell.value) for cell in header_cells]
    header_map = {normalize_header(name): index + 1 for index, name in enumerate(headers)}

    required = {
        "pageurl": "Page URL",
        "publicationtimeframe": "Publication timeframe",
        "topic": "Topic",
    }
    for key, label in required.items():
        if key not in header_map:
            raise ValueError(f"Missing required column: {label}")

    url_col = header_map["pageurl"]
    pub_col = header_map["publicationtimeframe"]
    topic_col = header_map["topic"]
    title_col = header_map.get("pagetitle")
    section_col = header_map.get("section")
    page_type_col = header_map.get("pagetype")

    publication_updates = 0
    topic_updates = 0
    rjina_updates = 0

    for row in range(2, ws.max_row + 1):
        url = extract_text(ws.cell(row, url_col).value)
        if not url:
            continue

        title = extract_text(ws.cell(row, title_col).value) if title_col else ""
        section = extract_text(ws.cell(row, section_col).value) if section_col else ""
        page_type = extract_text(ws.cell(row, page_type_col).value) if page_type_col else ""

        current_pub = extract_text(ws.cell(row, pub_col).value)
        if not current_pub:
            inferred_pub = infer_publication_timeframe(url, title)
            if inferred_pub:
                ws.cell(row, pub_col).value = inferred_pub
                publication_updates += 1

        current_topic = extract_text(ws.cell(row, topic_col).value)
        inferred_topic = infer_topic(current_topic, section, page_type)
        if inferred_topic and inferred_topic != current_topic:
            ws.cell(row, topic_col).value = inferred_topic
            topic_updates += 1

    if args.use_rjina:
        unresolved_rows: list[tuple[int, str, str]] = []
        for row in range(2, ws.max_row + 1):
            current_pub = extract_text(ws.cell(row, pub_col).value)
            if current_pub:
                continue
            url = extract_text(ws.cell(row, url_col).value)
            if not url:
                continue
            page_type = extract_text(ws.cell(row, page_type_col).value) if page_type_col else ""
            unresolved_rows.append((row, url, page_type))

        last_request_at = 0.0
        for row, url, page_type in unresolved_rows:
            elapsed = time.time() - last_request_at
            wait_time = args.rjina_min_interval_seconds - elapsed
            if wait_time > 0:
                time.sleep(wait_time)

            inferred_year, match_count = fetch_rjina_year(
                url, args.rjina_timeout_seconds, args.rjina_max_retries
            )
            last_request_at = time.time()
            if not inferred_year:
                continue

            # Avoid false positives from landing pages by requiring repeated dates,
            # except for explicit article pages.
            is_article = page_type.lower() == "press release / article page"
            if match_count < 2 and not is_article:
                continue

            ws.cell(row, pub_col).value = inferred_year
            publication_updates += 1
            rjina_updates += 1

    wb.save(args.output_file)
    print(f"Updated publication rows: {publication_updates}")
    print(f"Updated topic rows: {topic_updates}")
    if args.use_rjina:
        print(f"Updated publication rows via r.jina.ai: {rjina_updates}")
    print(f"Wrote: {args.output_file}")


if __name__ == "__main__":
    main()
