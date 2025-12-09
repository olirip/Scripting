#!/usr/bin/env python3
"""
Script to process Cloudinary error report from xlsx and convert to structured JSON.
"""

import json
from openpyxl import load_workbook
from pathlib import Path


def process_error_report(xlsx_path, json_output_path):
    """
    Process the xlsx error report and save as structured JSON.

    Structure:
    {
        "referrer": {
            "user_agent": {
                "code": {
                    "error": [requests]
                }
            }
        }
    }
    """
    # Load the workbook
    wb = load_workbook(xlsx_path)
    ws = wb.active

    # Get headers from first row
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    print(f"Headers found: {headers}")

    # Find column indices (case-insensitive search)
    def find_column_index(header_name):
        for idx, header in enumerate(headers):
            if header and header_name.lower() == header.lower():
                return idx
        return None

    referrer_idx = find_column_index('referrer')
    user_agent_idx = find_column_index('user_agent')
    if user_agent_idx is None:
        user_agent_idx = find_column_index('agent')
    code_idx = find_column_index('code')
    if code_idx is None:
        code_idx = find_column_index('status')
    error_idx = find_column_index('error')
    if error_idx is None:
        error_idx = find_column_index('message')
    request_idx = find_column_index('request')
    if request_idx is None:
        request_idx = find_column_index('url')
    if request_idx is None:
        request_idx = find_column_index('path')

    print(f"Column indices - Referrer: {referrer_idx}, User Agent: {user_agent_idx}, Code: {code_idx}, Error: {error_idx}, Request: {request_idx}")

    # Build the nested structure
    result = {}

    # Process each row (skip header)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        # Extract values
        referrer = row[referrer_idx] if referrer_idx is not None else None
        user_agent = row[user_agent_idx] if user_agent_idx is not None else None
        code = row[code_idx] if code_idx is not None else None
        error = row[error_idx] if error_idx is not None else None
        request = row[request_idx] if request_idx is not None else None

        # Handle empty or None referrer - use "none"
        if not referrer or referrer == "-":
            referrer = "none"
        else:
            referrer = str(referrer)

        # Convert other values to strings
        user_agent = str(user_agent) if user_agent else "unknown"
        code = str(code) if code else "unknown"
        error = str(error) if error else "unknown"
        request = str(request) if request else "unknown"

        # Clean up error message - remove "Resource not found - " prefix
        if error.startswith("Resource not found - "):
            error = error.replace("Resource not found - ", "", 1)

        # Replace /rolex-prod/ with https://media.rolex.com/
        request = request.replace("/rolex-prod/", "https://media.rolex.com/")

        # Build nested structure
        if referrer not in result:
            result[referrer] = {}

        if user_agent not in result[referrer]:
            result[referrer][user_agent] = {}

        if code not in result[referrer][user_agent]:
            result[referrer][user_agent][code] = {}

        if error not in result[referrer][user_agent][code]:
            result[referrer][user_agent][code][error] = []

        # Add request to the list (avoid duplicates)
        if request not in result[referrer][user_agent][code][error]:
            result[referrer][user_agent][code][error].append(request)

    # Save to JSON
    with open(json_output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    print(f"\nProcessed {row_idx - 1} rows")
    print(f"Output saved to: {json_output_path}")

    # Print some statistics
    total_referrers = len(result)
    total_requests = sum(
        len(requests)
        for referrer_data in result.values()
        for ua_data in referrer_data.values()
        for code_data in ua_data.values()
        for requests in code_data.values()
    )
    print(f"\nStatistics:")
    print(f"  Total unique referrers: {total_referrers}")
    print(f"  Total requests: {total_requests}")


if __name__ == "__main__":
    # Define file paths
    xlsx_file = "Cloudinary error report decembre.xlsx"
    json_file = "cloudinary_errors.json"

    # Check if xlsx file exists
    if not Path(xlsx_file).exists():
        print(f"Error: {xlsx_file} not found in current directory")
        exit(1)

    # Process the file
    process_error_report(xlsx_file, json_file)
